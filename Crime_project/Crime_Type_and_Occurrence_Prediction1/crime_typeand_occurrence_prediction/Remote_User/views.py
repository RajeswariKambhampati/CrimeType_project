from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
import datetime
import openpyxl


# Create your views here.
from Remote_User.models import ClientRegister_Model,Crime_details,Crime_type,detection_ratio,detection_accuracy

def login(request):
    if request.user.is_authenticated:
        return redirect('Add_DataSet_Details')

    if request.method == "POST" and 'submit1' in request.POST:
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if not username or not password:
            messages.error(request, 'Please provide both username and password.')
            return render(request, 'RUser/login.html')
        
        try:
            user = ClientRegister_Model.objects.get(username=username, password=password)
            request.session["userid"] = user.id
            request.session["username"] = user.username
            messages.success(request, f'Welcome back, {user.username}!')
            return redirect('Add_DataSet_Details')
        except ClientRegister_Model.DoesNotExist:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'RUser/login.html')

def Add_DataSet_Details(request):
    if 'userid' not in request.session:
        messages.error(request, 'Please login first.')
        return redirect('login')
    
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Please select an Excel file.')
            return render(request, 'RUser/Add_DataSet_Details.html', {})
        
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        #print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        #print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        #print(active_sheet)
        # reading a cell
        #print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        Crime_details.objects.all().delete()
        Crime_type.objects.all().delete()

        start_row = 2 if active_sheet.cell(1, 1).value in ('INCIDENT_NUMBER', 'OFFENSE_CODE', 'OFFENSE_DESCRIPTION') else 1
        for r in range(start_row, active_sheet.max_row + 1):
            Crime_details.objects.create(
        INCIDENT_NUMBER=active_sheet.cell(r, 1).value,
        OFFENSE_CODE=active_sheet.cell(r, 2).value,
        OFFENSE_CODE_GROUP=active_sheet.cell(r, 3).value,
        OFFENSE_DESCRIPTION=active_sheet.cell(r, 4).value,
        DISTRICT=active_sheet.cell(r, 5).value,
        REPORTING_AREA=active_sheet.cell(r, 6).value,
        OCCURRED_ON_DATE=active_sheet.cell(r, 7).value,
        YEAR=active_sheet.cell(r, 8).value,
        MONTH=active_sheet.cell(r, 9).value,
        DAY_OF_WEEK=active_sheet.cell(r, 10).value,
        Hour=active_sheet.cell(r, 11).value,
        UCR_PART=active_sheet.cell(r, 12).value,
        STREET=active_sheet.cell(r, 13).value,
        Lat=active_sheet.cell(r, 14).value,
        Long1=active_sheet.cell(r, 15).value,
        Location=active_sheet.cell(r, 16).value
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):
    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        
        # Validation
        if not all([username, email, password, confirm_password, phoneno, country, state, city]):
            messages.error(request, 'All fields are required.')
            return render(request, 'RUser/Register1.html')
        
        if password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'RUser/Register1.html')
        
        if len(password) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
            return render(request, 'RUser/Register1.html')
        
        if ClientRegister_Model.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'RUser/Register1.html')
        
        if ClientRegister_Model.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered.')
            return render(request, 'RUser/Register1.html')
        
        # Create user
        user = ClientRegister_Model.objects.create(
            username=username, 
            email=email, 
            password=password, 
            phoneno=phoneno,
            country=country, 
            state=state, 
            city=city
        )
        
        messages.success(request, 'Account created successfully! Please login.')
        return redirect('login')
    
    return render(request, 'RUser/Register1.html')

def logout_user(request):
    if 'userid' in request.session:
        del request.session['userid']
    if 'username' in request.session:
        del request.session['username']
    messages.success(request, 'Logged out successfully!')
    return redirect('index')

def ViewYourProfile(request):
    if 'userid' not in request.session:
        messages.error(request, 'Please login first.')
        return redirect('login')
    
    try:
        userid = request.session['userid']
        obj = ClientRegister_Model.objects.get(id=userid)
        return render(request, 'RUser/ViewYourProfile.html', {'object': obj})
    except ClientRegister_Model.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('login')


def Search_DataSets(request):
    if 'userid' not in request.session:
        messages.error(request, 'Please login first.')
        return redirect('login')
    
    if request.method == "POST":
        kword = request.POST.get('keyword')
        
        if not kword:
            messages.error(request, 'Please enter a search keyword.')
            return render(request, 'RUser/Search_DataSets.html')
        
        obj = Crime_details.objects.filter(Q(INCIDENT_NUMBER__icontains=kword))
        
        if not obj.exists():
            messages.info(request, 'No results found for your search.')
        
        return render(request, 'RUser/Search_DataSets.html', {'objs': obj, 'keyword': kword})
    
    return render(request, 'RUser/Search_DataSets.html')




